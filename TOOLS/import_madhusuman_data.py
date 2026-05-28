from __future__ import annotations

import json
import math
import re
import subprocess
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl


ROOT = Path(r"C:\HVAC_PRO_MSE")
SOURCE = Path(r"C:\Users\Administrator\Documents\MADHUSUMAN data")
STAGE = ROOT / "artifacts" / "madhusuman-import"
DATA = STAGE / "dummy-data" / "Dummy Data ( Harshal)"
OUT_DIR = STAGE / "generated"
SQL_FILE = OUT_DIR / "madhusuman_import.sql"
REPORT_FILE = OUT_DIR / "madhusuman_import_report.json"
SQLCMD = Path(r"C:\Program Files\Microsoft SQL Server\Client SDK\ODBC\170\Tools\Binn\SQLCMD.EXE")


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    text = str(value).replace("\xa0", " ").replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def title(value) -> str:
    text = clean(value)
    return text.title() if text.isupper() else text


def trunc(value, size: int) -> str:
    return clean(value)[:size]


def sql_text(value, size: int | None = None) -> str:
    text = clean(value)
    if size:
        text = text[:size]
    if not text:
        return "NULL"
    return "N'" + text.replace("'", "''") + "'"


def sql_text_required(value, fallback: str, size: int | None = None) -> str:
    text = clean(value) or fallback
    if size:
        text = text[:size]
    return "N'" + text.replace("'", "''") + "'"


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, bool):
        return Decimal("0")
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return Decimal("0")
        return Decimal(str(value))
    text = clean(value).replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def sql_decimal(value) -> str:
    amount = number(value).quantize(Decimal("0.01"))
    return str(amount)


def maybe_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 1000:
        try:
            return openpyxl.utils.datetime.from_excel(value).date()
        except Exception:
            return None
    text = clean(value)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def sql_date(value) -> str:
    dt = maybe_date(value)
    return "NULL" if dt is None else "'" + dt.isoformat() + "'"


def normalize_key(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean(value).upper())


def header_key(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean(value).upper())


def find_col(header: dict[str, int], *names: str, default: int | None = None) -> int | None:
    for name in names:
        key = header_key(name)
        if key in header:
            return header[key]
    return default


def cell(row, index: int | None, default=""):
    if index is None or index >= len(row):
        return default
    return row[index]


def rows(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            yield row
    finally:
        wb.close()


def parse_vendors() -> dict[str, dict]:
    vendors: dict[str, dict] = {}
    path = DATA / "Vendor details.xlsx"
    for sheet in ("Sheet1", "2025-26"):
        for idx, row in enumerate(rows(path, sheet), start=1):
            if idx == 1:
                continue
            name = trunc(row[1] if len(row) > 1 else "", 255)
            key = normalize_key(name)
            if not key:
                continue
            existing = vendors.setdefault(key, {"name": name})
            existing.update({
                "name": existing.get("name") or name,
                "address": existing.get("address") or trunc(row[2] if len(row) > 2 else "", 1000),
                "email": existing.get("email") or trunc(row[3] if len(row) > 3 else "", 255),
                "contact": existing.get("contact") or trunc(row[4] if len(row) > 4 else "", 100),
                "phone": existing.get("phone") or trunc(row[5] if len(row) > 5 else "", 20),
                "gst": existing.get("gst") or trunc(row[6] if len(row) > 6 else "", 50),
                "bank": existing.get("bank") or trunc(row[7] if len(row) > 7 else "", 30),
                "ifsc": existing.get("ifsc") or trunc(row[8] if len(row) > 8 else "", 11),
                "category": existing.get("category") or trunc(row[9] if len(row) > 9 else row[7] if len(row) > 7 else "", 100),
            })
    purchase_path = DATA / "Purchase Details April 2025 to March 2026.xlsx"
    for idx, row in enumerate(rows(purchase_path, "Masters"), start=1):
        if idx == 1:
            continue
        name = trunc(row[1] if len(row) > 1 else "", 255)
        key = normalize_key(name)
        if key:
            vendors.setdefault(key, {"name": name, "gst": trunc(row[2] if len(row) > 2 else "", 50)})
    return vendors


def parse_purchases() -> list[dict]:
    items: list[dict] = []
    path = DATA / "Purchase Details April 2025 to March 2026.xlsx"
    blank = 0
    for idx, row in enumerate(rows(path, "Purchases"), start=1):
        if idx <= 2:
            continue
        supplier = trunc(row[1] if len(row) > 1 else "", 255)
        inv_no = trunc(row[2] if len(row) > 2 else "", 100)
        if not supplier and not inv_no:
            blank += 1
            if blank > 50:
                break
            continue
        blank = 0
        if supplier.lower() == "cancelled":
            continue
        total = number(row[18] if len(row) > 18 else 0)
        if not supplier or total == 0:
            continue
        items.append({
            "supplier": supplier,
            "invoice_no": inv_no,
            "date": row[3] if len(row) > 3 else None,
            "subtotal": row[4] if len(row) > 4 else 0,
            "igst": number(row[5] if len(row) > 5 else 0) + number(row[6] if len(row) > 6 else 0),
            "cgst": sum(number(row[i]) for i in (7, 9, 11, 13, 15) if len(row) > i),
            "sgst": sum(number(row[i]) for i in (8, 10, 12, 14, 16) if len(row) > i),
            "total": total,
            "balance": row[19] if len(row) > 19 else 0,
            "gst": trunc(row[20] if len(row) > 20 else "", 50),
            "paid": row[22] if len(row) > 22 else 0,
            "site": trunc(row[23] if len(row) > 23 else "", 255),
            "due": row[25] if len(row) > 25 else None,
        })
    return items


def parse_stock() -> list[dict]:
    stock: list[dict] = []
    path = DATA / "Stock Summary.xlsx"
    blank = 0
    for idx, row in enumerate(rows(path, "Stock Summary"), start=1):
        if idx < 13:
            continue
        name = trunc(row[0] if len(row) > 0 else "", 255)
        qty = row[1] if len(row) > 1 else 0
        rate = row[2] if len(row) > 2 else 0
        value = row[3] if len(row) > 3 else 0
        if not name:
            blank += 1
            if blank > 50:
                break
            continue
        blank = 0
        if name.lower().startswith("grand total"):
            break
        if number(qty) == 0 and number(rate) == 0 and number(value) == 0:
            continue
        stock.append({
            "name": name,
            "qty": qty,
            "rate": rate if number(rate) != 0 else (number(value) / number(qty) if number(qty) != 0 else 0),
            "category": infer_category(name),
        })
    return stock


def infer_category(name: str) -> str:
    n = name.lower()
    if any(x in n for x in ("compressor", "condenser", "evaporator")):
        return "HVAC Major Parts"
    if any(x in n for x in ("pipe", "copper", "insulation", "duct")):
        return "Piping & Ducting"
    if any(x in n for x in ("gas", "refrigerant", "r-")):
        return "Refrigerants"
    if any(x in n for x in ("valve", "filter", "relay", "sensor", "capacitor")):
        return "Spares"
    return "General HVAC"


def parse_employees() -> dict[str, dict]:
    path = SOURCE / "employee list.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    employees: dict[str, dict] = {}
    try:
        for ws in wb.worksheets:
            if ws.max_column < 5:
                continue
            header_row = None
            header: dict[str, int] = {}
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                joined = " ".join(clean(v).upper() for v in row)
                if "NAME OF EMPLOYEE" in joined or "ID NO" in joined:
                    header_row = i
                    header = {header_key(v): idx for idx, v in enumerate(row) if clean(v)}
                    break
            start = (header_row + 1) if header_row else 3
            code_col = find_col(header, "ID No", default=1)
            esic_col = find_col(header, "ESIC NUMBER", "INSURANCE NUMBER", default=2)
            uan_col = find_col(header, "UAN NUMBER", default=3)
            name_col = find_col(header, "NAME OF EMPLOYEE", default=4)
            doj_col = find_col(header, "D.O.J.", "DOJ", default=6)
            dob_col = find_col(header, "D.O.B", "DOB", default=7)
            marital_col = find_col(header, "MARITAL STATUS", default=8)
            address_col = find_col(header, "ADDRESS", default=9)
            pan_col = find_col(header, "PAN NUMBER", default=13)
            aadhaar_col = find_col(header, "AADHAR CARD NUMBER", "AADHAAR CARD NUMBER", default=14)
            bank_col = find_col(header, "BANK A/C NUMBER", "BANK AC NUMBER", default=15)
            ifsc_col = find_col(header, "IFSC CODE", default=16)
            phone_col = find_col(header, "MOBILE NO", "MOBILE  NO", default=17)
            basic_col = find_col(header, "Basic", default=18)
            gross_col = find_col(header, "Gross", default=21)
            nature_col = find_col(header, "Nature of WORK", default=23)
            site_col = find_col(header, "CLIENT", default=24)
            epf_col = find_col(header, "EPF No", default=25)
            status_col = find_col(header, "Status", default=26)
            designation_col = find_col(header, "Designation", default=28)
            blank = 0
            for row in ws.iter_rows(min_row=start, values_only=True):
                code = trunc(cell(row, code_col), 50)
                name = trunc(title(cell(row, name_col)), 255)
                if not code and not name:
                    blank += 1
                    if blank > 200:
                        break
                    continue
                blank = 0
                if not code or not name or code.lower() == "blank" or name.lower() == "blank":
                    continue
                key = normalize_key(code)
                if key in employees:
                    continue
                employees[key] = {
                    "code": code,
                    "name": name,
                    "esic": trunc(cell(row, esic_col), 50),
                    "uan": trunc(cell(row, uan_col), 12),
                    "doj": cell(row, doj_col, None),
                    "dob": cell(row, dob_col, None),
                    "marital": trunc(cell(row, marital_col), 50),
                    "address": trunc(cell(row, address_col), 2000),
                    "pan": trunc(cell(row, pan_col), 10),
                    "aadhaar": trunc(cell(row, aadhaar_col), 12),
                    "bank": trunc(cell(row, bank_col), 20),
                    "ifsc": trunc(cell(row, ifsc_col), 11),
                    "phone": trunc(cell(row, phone_col), 15),
                    "basic": cell(row, basic_col, 0),
                    "gross": cell(row, gross_col, 0),
                    "nature": trunc(cell(row, nature_col), 50),
                    "site": trunc(cell(row, site_col) or ws.title, 255),
                    "epf": trunc(cell(row, epf_col), 50),
                    "status": trunc(cell(row, status_col) or "Active", 50),
                    "designation": trunc(cell(row, designation_col), 255),
                }
    finally:
        wb.close()
    return employees


def parse_clients_and_invoices() -> tuple[dict[str, dict], list[dict]]:
    clients: dict[str, dict] = {}
    invoices: list[dict] = []
    path = STAGE / "sales-tax-invoices" / "Sales Tax Invoice from April 2025 to March 2026.xlsx"
    for idx, row in enumerate(rows(path, "GSTIN Master"), start=1):
        if idx == 1:
            continue
        name = trunc(row[1] if len(row) > 1 else "", 255)
        key = normalize_key(name)
        if key:
            clients.setdefault(key, {"name": name, "gst": trunc(row[2] if len(row) > 2 else "", 50)})
    blank = 0
    for idx, row in enumerate(rows(path, "Sales"), start=1):
        if idx <= 2:
            continue
        customer = trunc(row[1] if len(row) > 1 else "", 255)
        inv_no = trunc(row[3] if len(row) > 3 else "", 100)
        if not customer and not inv_no:
            blank += 1
            if blank > 50:
                break
            continue
        blank = 0
        if not customer or not inv_no:
            continue
        clients.setdefault(normalize_key(customer), {"name": customer})
        subtotal = number(row[7] if len(row) > 7 else 0)
        igst = number(row[8] if len(row) > 8 else 0) + number(row[9] if len(row) > 9 else 0)
        cgst = sum(number(row[i]) for i in (10, 12, 14) if len(row) > i)
        sgst = sum(number(row[i]) for i in (11, 13, 15) if len(row) > i)
        total = number(row[16] if len(row) > 16 else 0)
        tds = number(row[18] if len(row) > 18 else 0)
        adv = number(row[19] if len(row) > 19 else 0)
        deductions = number(row[20] if len(row) > 20 else 0)
        credit = number(row[21] if len(row) > 21 else 0)
        settled = credit + tds + adv + deductions
        remark = trunc(row[22] if len(row) > 22 else "", 100)
        paid = total if remark.lower() == "paid" and total > 0 else settled
        invoices.append({
            "customer": customer,
            "date": row[2] if len(row) > 2 else None,
            "invoice_no": inv_no,
            "po": trunc(row[4] if len(row) > 4 else "", 100),
            "po_date": row[5] if len(row) > 5 else None,
            "description": trunc(row[6] if len(row) > 6 else "Imported sales invoice", 1000),
            "subtotal": subtotal,
            "igst": igst,
            "cgst": cgst,
            "sgst": sgst,
            "tax": igst + cgst + sgst,
            "total": total if total > 0 else subtotal + igst + cgst + sgst,
            "paid": paid,
            "payment_date": row[17] if len(row) > 17 else None,
            "remarks": remark,
            "due": row[24] if len(row) > 24 else None,
        })
    return clients, invoices


def po_number(invoice_no: str, index: int) -> str:
    safe = re.sub(r"[^A-Za-z0-9]+", "-", clean(invoice_no)).strip("-")
    return ("PO-" + safe)[:100] if safe else f"PO-IMP-{index:04d}"


def gst_percent(subtotal: Decimal, tax: Decimal) -> str:
    if subtotal <= 0 or tax <= 0:
        return "0"
    return str((tax * Decimal("100") / subtotal).quantize(Decimal("0.01")))


def purchase_tax_rate(amount: Decimal) -> str:
    return "9.00" if amount > 0 else "0.00"


def purchase_igst_rate(amount: Decimal) -> str:
    return "18.00" if amount > 0 else "0.00"


def build_sql():
    vendors = parse_vendors()
    purchases = parse_purchases()
    stock = parse_stock()
    employees = parse_employees()
    clients, invoices = parse_clients_and_invoices()

    lines: list[str] = [
        "SET XACT_ABORT ON;",
        "BEGIN TRY",
        "BEGIN TRANSACTION;",
        "DELETE FROM Employees WHERE EmployeeCode = N'MSE000' AND Name = N'Blank';",
    ]

    for v in vendors.values():
        lines.append(f"""
IF NOT EXISTS (SELECT 1 FROM Vendors WHERE UPPER(LTRIM(RTRIM(VendorName))) = UPPER(LTRIM(RTRIM({sql_text_required(v.get('name'), 'Unknown Vendor', 255)}))))
BEGIN
    INSERT INTO Vendors (VendorName, GSTNumber, PANNumber, Phone, Email, Address, City, Category, IsActive, BankAccountNumber, BankIFSC, BankAccountName, VendorType, Notes)
    VALUES ({sql_text_required(v.get('name'), 'Unknown Vendor', 255)}, {sql_text(v.get('gst'), 50)}, NULL, {sql_text(v.get('phone'), 20)}, {sql_text(v.get('email'), 255)}, {sql_text(v.get('address'), 1000)}, NULL, {sql_text(v.get('category'), 100)}, 1, {sql_text(v.get('bank'), 30)}, {sql_text(v.get('ifsc'), 11)}, {sql_text(v.get('contact'), 100)}, N'Supplier', N'Imported from Madhusuman vendor workbook');
END;""")

    for c in clients.values():
        lines.append(f"""
IF NOT EXISTS (SELECT 1 FROM B2BClients WHERE UPPER(LTRIM(RTRIM(CompanyName))) = UPPER(LTRIM(RTRIM({sql_text_required(c.get('name'), 'Unknown Client', 255)}))))
BEGIN
    INSERT INTO B2BClients (CompanyName, IndustryType, GSTNumber, PaymentTermsDays, CreditLimit, IsActive, RelationshipStage, Notes)
    VALUES ({sql_text_required(c.get('name'), 'Unknown Client', 255)}, N'HVAC / Facility Services', {sql_text(c.get('gst'), 50)}, 30, 0, 1, N'Customer', N'Imported from Madhusuman sales invoice workbook');
END;""")

    for e in employees.values():
        aadhaar = clean(e.get("aadhaar"))
        last4 = aadhaar[-4:] if len(aadhaar) >= 4 else ""
        lines.append(f"""
IF NOT EXISTS (SELECT 1 FROM Employees WHERE EmployeeCode = {sql_text_required(e.get('code'), 'UNKNOWN', 50)})
BEGIN
    INSERT INTO Employees
        (EmployeeCode, Name, Designation, Department, ClientSite, Phone, JoiningDate, Status,
         DateOfBirth, MaritalStatus, Address, PAN, PANNumber, ESICNumber, UANNumber, UAN, EPFNumber,
         BankAccount, BankAccountNumber, IFSCCode, BankIFSC, NatureOfWork, BasicSalary, GrossSalary,
         AadhaarNumber, AadhaarLast4, DateOfJoining, EmploymentType, EPFApplicable, ESIApplicable, PTApplicable,
         CreatedByName)
    VALUES
        ({sql_text_required(e.get('code'), 'UNKNOWN', 50)}, {sql_text_required(e.get('name'), 'Unknown Employee', 255)}, {sql_text(e.get('designation'), 255)}, {sql_text(e.get('nature'), 100)}, {sql_text(e.get('site'), 255)}, {sql_text(e.get('phone'), 20)}, {sql_date(e.get('doj'))}, {sql_text_required(e.get('status'), 'Active', 50)},
         {sql_date(e.get('dob'))}, {sql_text(e.get('marital'), 50)}, {sql_text(e.get('address'), 2000)}, {sql_text(e.get('pan'), 20)}, {sql_text(e.get('pan'), 10)}, {sql_text(e.get('esic'), 50)}, {sql_text(e.get('uan'), 50)}, {sql_text(e.get('uan'), 12)}, {sql_text(e.get('epf'), 50)},
         {sql_text(e.get('bank'), 50)}, {sql_text(e.get('bank'), 20)}, {sql_text(e.get('ifsc'), 20)}, {sql_text(e.get('ifsc'), 11)}, {sql_text(e.get('nature'), 50)}, {sql_decimal(e.get('basic'))}, {sql_decimal(e.get('gross'))},
         {sql_text(aadhaar, 12)}, {sql_text(last4, 4)}, {sql_date(e.get('doj'))}, N'Permanent', CASE WHEN {sql_text(e.get('epf'), 50)} IS NULL THEN 0 ELSE 1 END, CASE WHEN {sql_text(e.get('esic'), 50)} IS NULL THEN 0 ELSE 1 END, 0,
         N'Madhusuman import');
END
ELSE
BEGIN
    UPDATE Employees SET
        Name = {sql_text_required(e.get('name'), 'Unknown Employee', 255)},
        Designation = {sql_text(e.get('designation'), 255)},
        Department = {sql_text(e.get('nature'), 100)},
        ClientSite = {sql_text(e.get('site'), 255)},
        Phone = {sql_text(e.get('phone'), 20)},
        JoiningDate = {sql_date(e.get('doj'))},
        Status = {sql_text_required(e.get('status'), 'Active', 50)},
        DateOfBirth = {sql_date(e.get('dob'))},
        MaritalStatus = {sql_text(e.get('marital'), 50)},
        Address = {sql_text(e.get('address'), 2000)},
        PAN = {sql_text(e.get('pan'), 20)},
        PANNumber = {sql_text(e.get('pan'), 10)},
        ESICNumber = {sql_text(e.get('esic'), 50)},
        UANNumber = {sql_text(e.get('uan'), 50)},
        UAN = {sql_text(e.get('uan'), 12)},
        EPFNumber = {sql_text(e.get('epf'), 50)},
        BankAccount = {sql_text(e.get('bank'), 50)},
        BankAccountNumber = {sql_text(e.get('bank'), 20)},
        IFSCCode = {sql_text(e.get('ifsc'), 20)},
        BankIFSC = {sql_text(e.get('ifsc'), 11)},
        NatureOfWork = {sql_text(e.get('nature'), 50)},
        BasicSalary = {sql_decimal(e.get('basic'))},
        GrossSalary = {sql_decimal(e.get('gross'))},
        AadhaarNumber = {sql_text(aadhaar, 12)},
        AadhaarLast4 = {sql_text(last4, 4)},
        DateOfJoining = {sql_date(e.get('doj'))},
        EPFApplicable = CASE WHEN {sql_text(e.get('epf'), 50)} IS NULL THEN 0 ELSE 1 END,
        ESIApplicable = CASE WHEN {sql_text(e.get('esic'), 50)} IS NULL THEN 0 ELSE 1 END,
        ModifiedByName = N'Madhusuman import',
        ModifiedDate = GETDATE()
    WHERE EmployeeCode = {sql_text_required(e.get('code'), 'UNKNOWN', 50)};
END;""")

    for s in stock:
        lines.append(f"""
IF NOT EXISTS (SELECT 1 FROM StockItems WHERE UPPER(LTRIM(RTRIM(ItemName))) = UPPER(LTRIM(RTRIM({sql_text_required(s.get('name'), 'Unknown Item', 255)}))))
BEGIN
    INSERT INTO StockItems (ItemName, Category, CurrentStock, Unit, LastPurchaseRate, ReorderLevel)
    VALUES ({sql_text_required(s.get('name'), 'Unknown Item', 255)}, {sql_text(s.get('category'), 100)}, {sql_decimal(s.get('qty'))}, N'Nos', {sql_decimal(s.get('rate'))}, CASE WHEN ABS({sql_decimal(s.get('qty'))}) > 10 THEN 5 ELSE 1 END);
END;""")

    for i, p in enumerate(purchases, start=1):
        ponum = po_number(p["invoice_no"], i)
        notes = "Imported from Madhusuman purchase workbook"
        if p.get("site"):
            notes += " | Site: " + p["site"]
        if number(p.get("balance")) > 0:
            notes += " | Balance: " + str(number(p.get("balance")).quantize(Decimal("0.01")))
        status = "Paid" if number(p.get("paid")) >= number(p.get("total")) and number(p.get("total")) > 0 else ("Partial" if number(p.get("paid")) > 0 else "Pending")
        lines.append(f"""
IF NOT EXISTS (SELECT 1 FROM PurchaseOrders WHERE VendorInvoiceNumber = {sql_text(p.get('invoice_no'), 100)} AND VendorID = (SELECT TOP 1 VendorID FROM Vendors WHERE UPPER(LTRIM(RTRIM(VendorName))) = UPPER(LTRIM(RTRIM({sql_text_required(p.get('supplier'), 'Unknown Vendor', 255)})))))
BEGIN
    DECLARE @VendorId_{i} INT = (SELECT TOP 1 VendorID FROM Vendors WHERE UPPER(LTRIM(RTRIM(VendorName))) = UPPER(LTRIM(RTRIM({sql_text_required(p.get('supplier'), 'Unknown Vendor', 255)}))));
    IF @VendorId_{i} IS NULL
    BEGIN
        INSERT INTO Vendors (VendorName, GSTNumber, IsActive, VendorType, Notes) VALUES ({sql_text_required(p.get('supplier'), 'Unknown Vendor', 255)}, {sql_text(p.get('gst'), 50)}, 1, N'Supplier', N'Created during Madhusuman purchase import');
        SET @VendorId_{i} = SCOPE_IDENTITY();
    END;
    INSERT INTO PurchaseOrders (VendorID, PONumber, PODate, PayByDate, VendorInvoiceNumber, TotalAmount, PaidAmount, Status, Notes, CreatedByName)
    VALUES (@VendorId_{i}, {sql_text_required(ponum, f'PO-IMP-{i:04d}', 100)}, COALESCE({sql_date(p.get('date'))}, GETDATE()), {sql_date(p.get('due'))}, {sql_text(p.get('invoice_no'), 100)}, {sql_decimal(p.get('total'))}, {sql_decimal(p.get('paid'))}, N'{status}', {sql_text(notes, 1000)}, N'Madhusuman import');
    DECLARE @POID_{i} INT = SCOPE_IDENTITY();
    INSERT INTO PurchaseLineItems (POID, Description, Quantity, Rate, Amount, GSTRate, CGSTRate, SGSTRate, IGSTRate, JobLink, UOM)
    VALUES (@POID_{i}, {sql_text_required('Imported purchase invoice ' + clean(p.get('invoice_no')), 'Imported purchase invoice', 255)}, 1, {sql_decimal(p.get('subtotal'))}, {sql_decimal(p.get('total'))}, {gst_percent(number(p.get('subtotal')), number(p.get('total')) - number(p.get('subtotal')))}, {purchase_tax_rate(number(p.get('cgst')))}, {purchase_tax_rate(number(p.get('sgst')))}, {purchase_igst_rate(number(p.get('igst')))}, N'General', N'Nos');
END;""")

    for i, inv in enumerate(invoices, start=1):
        status = "Cancelled" if "cancel" in clean(inv.get("remarks")).lower() else ("Paid" if inv["paid"] >= inv["total"] and inv["total"] > 0 else ("Partial" if inv["paid"] > 0 else "Pending"))
        balance = max(Decimal("0"), inv["total"] - inv["paid"])
        gst_mode = "IGST" if inv["igst"] > 0 else "CGST_SGST"
        notes = "Imported from Madhusuman sales tax invoice workbook"
        if inv.get("remarks"):
            notes += " | Remark: " + inv["remarks"]
        lines.append(f"""
IF NOT EXISTS (SELECT 1 FROM Invoices WHERE InvoiceNumber = {sql_text_required(inv.get('invoice_no'), f'INV-IMP-{i:04d}', 100)})
BEGIN
    DECLARE @ClientId_{i} INT = (SELECT TOP 1 ClientID FROM B2BClients WHERE UPPER(LTRIM(RTRIM(CompanyName))) = UPPER(LTRIM(RTRIM({sql_text_required(inv.get('customer'), 'Unknown Client', 255)}))));
    IF @ClientId_{i} IS NULL
    BEGIN
        INSERT INTO B2BClients (CompanyName, IndustryType, PaymentTermsDays, CreditLimit, IsActive, RelationshipStage, Notes) VALUES ({sql_text_required(inv.get('customer'), 'Unknown Client', 255)}, N'HVAC / Facility Services', 30, 0, 1, N'Customer', N'Created during Madhusuman invoice import');
        SET @ClientId_{i} = SCOPE_IDENTITY();
    END;
    INSERT INTO Invoices
        (ClientID, InvoiceNumber, InvoiceDate, DueDate, SubTotal, TaxAmount, TotalAmount, PaidAmount, PaymentStatus, PaymentDate,
         GSTPercent, BalanceDue, Notes, InvoiceTitle, Subject, PONumber, PODate, GSTMode, CGSTAmount, SGSTAmount, IGSTAmount,
         CreatedByName)
    VALUES
        (@ClientId_{i}, {sql_text_required(inv.get('invoice_no'), f'INV-IMP-{i:04d}', 100)}, {sql_date(inv.get('date'))}, {sql_date(inv.get('due'))}, {sql_decimal(inv.get('subtotal'))}, {sql_decimal(inv.get('tax'))}, {sql_decimal(inv.get('total'))}, {str(inv['paid'].quantize(Decimal('0.01')))}, N'{status}', {sql_date(inv.get('payment_date'))},
         {gst_percent(inv['subtotal'], inv['tax'])}, {str(balance.quantize(Decimal('0.01')))}, {sql_text(notes, 1000)}, N'TAX INVOICE', {sql_text(inv.get('description'), 500)}, {sql_text(inv.get('po'), 100)}, {sql_date(inv.get('po_date'))}, N'{gst_mode}', {sql_decimal(inv.get('cgst'))}, {sql_decimal(inv.get('sgst'))}, {sql_decimal(inv.get('igst'))},
         N'Madhusuman import');
    DECLARE @InvoiceId_{i} INT = SCOPE_IDENTITY();
    INSERT INTO InvoiceLineItems (InvoiceID, Description, Unit, Quantity, Rate, Amount, Category, GSTPercent, TaxAmount, TaxType, IsBillable)
    VALUES (@InvoiceId_{i}, {sql_text_required(inv.get('description'), 'Imported sales invoice', 1000)}, N'Nos', 1, {sql_decimal(inv.get('subtotal'))}, {sql_decimal(inv.get('subtotal'))}, N'Service', {gst_percent(inv['subtotal'], inv['tax'])}, {sql_decimal(inv.get('tax'))}, N'Taxable', 1);
END;""")

    lines += [
        "COMMIT TRANSACTION;",
        "END TRY",
        "BEGIN CATCH",
        "IF @@TRANCOUNT > 0 ROLLBACK TRANSACTION;",
        "THROW;",
        "END CATCH;",
    ]

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    SQL_FILE.write_text("\n".join(lines), encoding="utf-8")
    REPORT_FILE.write_text(json.dumps({
        "vendors": len(vendors),
        "purchases": len(purchases),
        "stock_items": len(stock),
        "employees": len(employees),
        "clients": len(clients),
        "invoices": len(invoices),
        "sql_file": str(SQL_FILE),
    }, indent=2), encoding="utf-8")


def main():
    build_sql()
    subprocess.run([str(SQLCMD), "-S", r".\SQLEXPRESS", "-E", "-d", "HVAC_PRO", "-i", str(SQL_FILE), "-b"], check=True)
    print(REPORT_FILE.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
