from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


OUT_DIR = Path(r"C:\HVAC_PRO_MSE\TEST_DATA\quotations")
OUT_FILE = OUT_DIR / "ServoERP_Quotation_Test_Import.xlsx"


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotations"
    headers = [
        "QuotationNumber",
        "QuotationDate",
        "ClientName",
        "SiteName",
        "Description",
        "Amount",
        "Status",
        "ValidUntil",
    ]
    rows = [
        ["QT-TEST-2026-001", "2026-05-26", "Atul", "", "Test quotation for compressor servicing and preventive maintenance", 18500, "Draft", "2026-06-10"],
        ["QT-TEST-2026-002", "2026-05-26", "Zydus", "", "Test quotation for AHU inspection, filter replacement, and labour", 42750, "Submitted", "2026-06-12"],
        ["QT-TEST-2026-003", "2026-05-26", "Aarti", "", "Test quotation for split AC repair materials and technician visit", 9600, "Approved", "2026-06-15"],
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)

    fill = PatternFill("solid", fgColor="DBEAFE")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    widths = [22, 18, 24, 18, 72, 14, 16, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    main()
